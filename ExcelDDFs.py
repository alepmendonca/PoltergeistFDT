import datetime
import math
import os
import shutil
import openpyxl
import pandas as pd
import numpy as np
import re
import pywintypes
import win32com.client
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment
from copy import copy
from pandas import Timestamp
from pathlib import Path
from pandas._libs.missing import NAType

import Audit
import GeneralConfiguration
import GeneralFunctions
import MDBReader
import PDFExtractor
from ConfigFiles import Analysis, Infraction
from GeneralFunctions import logger
from SQLReader import SQLReader
from WebScraper import SeleniumWebScraper


class ExcelArrazoadoException(Exception):
    pass


class ExcelArrazoadoCriticalException(ExcelArrazoadoException):
    pass


class ExcelArrazoadoIncompletoException(ExcelArrazoadoException):
    pass


class ExcelArrazoadoAbaInexistenteException(ExcelArrazoadoException):
    pass


def _dataframe_to_rows(df: pd.DataFrame, header=True):
    """
    Copiado e adaptado do openpyxl, mas corrigindo comportamento para Int64
    Convert a Pandas dataframe into something suitable for passing into a worksheet.
    If index is True then the index will be included, starting one row below the header.
    If header is True then column headers will be included starting one column to the right.
    Formatting should be done by client code.
    """
    blocks = df._data.blocks
    ncols = sum(b.shape[0] for b in blocks)
    data = [None] * ncols

    for b in blocks:
        values = b.values

        if b.dtype.type == np.datetime64:
            values = np.array([Timestamp(v) for v in values.ravel()])
            values = values.reshape(b.shape)
        elif b.dtype.name == 'Int64':
            values = np.array([v for v in values.ravel()])
            values = values.reshape(b.shape)

        result = values.tolist()

        for col_loc, col in zip(b.mgr_locs, result):
            data[col_loc] = col

    if header:
        rows = [list(df.columns.values)]
        for row in rows:
            n = []
            for v in row:
                if isinstance(v, np.datetime64):
                    v = Timestamp(v)
                n.append(v)
            row = n
            yield row

    expanded = ([v] for v in df.index)
    for idx, v in enumerate(expanded):
        row = [data[j][idx] if not isinstance(data[j][idx], NAType) else None for j in range(ncols)]
        yield row


def refresh_planilha(path: Path):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.ScreenUpdating = False
    excel.DisplayAlerts = False
    excel.EnableEvents = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(path))
        wb.RefreshAll()
        excel.CalculateUntilAsyncqueriesDone()
        wb.Save()
    except pywintypes.com_error as comerror:
        raise ExcelArrazoadoCriticalException(f'Falha no Excel: {comerror.excepinfo[2]}')
    finally:
        if wb:
            wb.Close()
        excel.Quit()


def print_workbook_as_pdf(wb_path: Path, pdf: Path, sheet_number: int = None, header=False) -> Path:
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.ScreenUpdating = False
    excel.DisplayAlerts = False
    excel.EnableEvents = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(wb_path))
        # Array em COM começa em 1, enquanto arrays em Python começam em 0
        if sheet_number is not None:
            sheets = [sheet_number + 1]
        else:
            sheets = list(range(1, int(wb.Worksheets.Count) + 1))
        # Serve para corrigir o uso das definições de PageSetup no COM
        for idx in sheets:
            ws = wb.Worksheets[idx - 1]
            if header:
                ws.PageSetup.CenterHeader = f'Arquivo {wb_path.name} - Planilha {ws.Name}'
                # Landscape se mais de 10 colunas
                if len(ws.Columns) > 10:
                    ws.PageSetup.Orientation = 2
                else:
                    ws.PageSetup.Orientation = 1
            for name in ws.Names:
                if name.Name.endswith('!Print_Area'):
                    ws.PageSetup.PrintArea = name.RefersTo
        pdf.unlink(missing_ok=True)
        # assinatura do método ExportedAsFixedFormat disponível em
        # https://docs.microsoft.com/pt-br/office/vba/api/excel.workbook.exportasfixedformat
        if sheet_number is not None:
            ws = wb.Worksheets[sheet_number]
            ws.ExportAsFixedFormat(0, str(pdf.absolute()), 0, True, False)
        else:
            wb.Worksheets(sheets).Select()
            wb.ActiveSheet.ExportAsFixedFormat(0, str(pdf.absolute()), 0, True, False)
        return pdf
    except pywintypes.com_error as comerror:
        raise ExcelArrazoadoCriticalException(f'Falha no Excel: {comerror.excepinfo[2]}')
    finally:
        if wb:
            wb.Close()
        excel.Quit()


def glosas_item_path(item: int) -> Path:
    return Audit.get_current_audit().findings_path() / f'Glosa - Item {item}.xlsx'


class ExcelDDFs:
    def __init__(self):
        self.main_path = Audit.get_current_audit().path()
        empresa = GeneralFunctions.get_default_name_for_business(Audit.get_current_audit().empresa)

        Audit.get_current_audit().findings_path().mkdir(exist_ok=True)
        self.planilha_path = Audit.get_current_audit().findings_path() / f'Arrazoado - {empresa}.xlsm'
        self._workbook = None
        if not self.planilha_path.is_file():
            try:
                shutil.copyfile(os.path.join('resources', 'template.xlsm'), self.planilha_path)
                ws = self.workbook()['EFDs']
                ws['A4'].value = f"{GeneralConfiguration.get().drt_nome} - {GeneralConfiguration.get().drt_sigla}"
                ws['A5'].value = f"NÚCLEO DE FISCALIZAÇÃO - NF {GeneralConfiguration.get().nucleo_fiscal()} - " \
                                 f"EQUIPE FISCAL {GeneralConfiguration.get().equipe_fiscal}"
                ws['A8'].value = f'{Audit.get_current_audit().empresa} - IE {Audit.get_current_audit().ie}'
                self.salva_planilha()
            except Exception as e:
                logger.exception('Ocorreu uma falha ao copiar a planilha template para a pasta da fiscalização')
                try:
                    self.planilha_path.unlink(missing_ok=True)
                except Exception:
                    pass
                raise e
        self._planilha: dict[str, pd.DataFrame] = {}
        self.creditos_path = Audit.get_current_audit().findings_path() / 'Glosa de Créditos.xlsx'
        self.quadro3_path = Audit.get_current_audit().aiim_path() / 'Quadro 3.xlsm'
        self.operacoes_csv = Audit.get_current_audit().reports_path() / 'valor_operacoes.csv'
        self._template_row = 0

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.clear_cache()

    def _get_template_row(self) -> int:
        if self._template_row == 0:
            self._template_row = self.workbook()['template'].max_row
        return self._template_row

    def workbook(self) -> Workbook:
        if self._workbook is None:
            logger.info('Abrindo planilha da auditoria...')
            self._workbook = openpyxl.load_workbook(self.planilha_path, keep_vba=True)
        return self._workbook

    def planilha(self, sheet_name: str) -> pd.DataFrame:
        if not self._planilha:
            logger.info('Recalculando fórmulas da planilha...')
            refresh_planilha(self.planilha_path)
        if self._planilha.get(sheet_name) is None:
            try:
                self._planilha[sheet_name] = pd.read_excel(self.planilha_path, sheet_name=sheet_name,
                                                           header=self._get_template_row() - 1)
            except ValueError as ex:
                if str(ex) == f"Worksheet named '{sheet_name}' not found":
                    raise ExcelArrazoadoAbaInexistenteException(f'Aba {sheet_name} não existe mais na planilha!')
                else:
                    raise ex
            # remove colunas que não sejam com título em string
            for key in self._planilha[sheet_name].keys():
                if not isinstance(key, str) or key.startswith('Unnamed'):
                    self._planilha[sheet_name].drop(key, inplace=True, axis=1)
        return self._planilha[sheet_name]

    def get_sheet_as_df(self, sheet_name) -> pd.DataFrame:
        df = self.planilha(sheet_name)
        if 'Chave' in df.keys():
            df = df[df['Chave'] >= '1']
        aggregator = [i for i, x in enumerate(df.keys()) if isinstance(x, str) and x.lower() in ('mês', 'dia')]
        if aggregator:
            df = df[df[df.keys()[0]].map(lambda v: not isinstance(v, str) or not v.upper().startswith('TOTAL'))]
            df = df.iloc[:, 0:aggregator[0]]
        df = df.convert_dtypes()
        # converte tipo de colunas datas, para uso em comparações
        new_types = {}
        for idx in range(0, len(df.keys())):
            t = df.dtypes.tolist()[idx]
            coluna = df.keys()[idx]
            if t == np.object_ and any(df[coluna].map(
                    lambda v: isinstance(v, datetime.datetime) or isinstance(v, datetime.date) or
                              isinstance(v, Timestamp)
            )):
                new_types[coluna] = np.datetime64
        if new_types:
            df = df.astype(new_types)
        return df

    def create_creditos_sheet(self, saldos: pd.DataFrame):
        wb = None
        if not self.creditos_path.is_file():
            try:
                logger.info('Criando planilha de glosa de créditos com saldos de GIA da CFICMS')
                shutil.copyfile(os.path.join('resources', 'Glosa de Créditos.xlsx'),
                                self.creditos_path)
                wb = openpyxl.load_workbook(self.creditos_path)
                ws = wb['Dados']
                periodos_rpa = Audit.get_current_audit().get_periodos_da_fiscalizacao()
                if len(periodos_rpa) == 0:
                    raise ExcelArrazoadoCriticalException('Contribuinte não possui nenhum histórico de RPA no arquivo!'
                                                          '\nPara gerar autuação sobre créditos, é necessário verificar'
                                                          ' os saldos de GIA.\n'
                                                          'Atualize os dados da fiscalizada e tente novamente.')
                mes_referencia = GeneralFunctions.last_day_of_month(periodos_rpa[0][0])
                ws['A2'].value = mes_referencia

                linha_saldo = 2
                for _, referencia in saldos.iterrows():
                    if mes_referencia != referencia['referencia'].date():
                        raise ExcelArrazoadoCriticalException(f'Não foi encontrada na Conta Fiscal ICMS saldo para '
                                                              f'referência {mes_referencia}! '
                                                              f'Talvez precise editar arquivo cficms.json '
                                                              f'manualmente...')
                    ws.cell(linha_saldo, 8).value = referencia['saldo']
                    linha_saldo += 1
                    mes_referencia = GeneralFunctions.last_day_of_month(mes_referencia + datetime.timedelta(days=1))
                self.salva_planilha(wb, self.creditos_path)
            except Exception as e:
                logger.exception('Ocorreu uma falha ao copiar a planilha de créditos '
                                 'template para a pasta da fiscalização')
                try:
                    self.creditos_path.unlink(missing_ok=True)
                except Exception:
                    pass
                raise e
            finally:
                if wb is not None:
                    wb.close()

    def generate_creditos_ddf(self, ddf: pd.DataFrame, alinea: str, item: int) -> pd.DataFrame:
        wb = None
        try:
            logger.info('Preenchendo planilha de glosa de créditos com valores a glosar')
            glosas_path = glosas_item_path(item)
            shutil.copyfile(self.creditos_path, glosas_path)
            wb = openpyxl.load_workbook(glosas_path)
            ws = wb['Dados']
            mes_inicial = ws['A2'].value
            linha_inicial = 2
            for _, linha in ddf.iterrows():
                if not math.isnan(linha['valor']):
                    mes = linha['referencia']
                    diferenca_referencias = relativedelta(mes, mes_inicial)
                    meses = diferenca_referencias.years * 12 + diferenca_referencias.months
                    ws.cell(meses + linha_inicial, 6).value = linha['valor']
            self.salva_planilha(wb, glosas_path)

            glosas = pd.read_excel(glosas_path, sheet_name='Subitens')
            glosas = glosas[glosas['subitem'] > 0]
            glosas = glosas.iloc[:, [1, 11, 17, 18, 19]]
            glosas['referencia'] = glosas.iloc[:, 0].map(lambda dt: GeneralFunctions.last_day_of_month(dt)) \
                .astype(np.datetime64)
            resultado = ddf.merge(glosas, on='referencia')
            if 'valor_basico' in resultado.columns:
                resultado = resultado.iloc[:, [1, 3, 7, 8, 9, 10]]
                resultado.columns = ['valor_basico', 'referencia', 'valor', 'dci', 'dij', 'dcm']
                resultado['valor_basico'] = resultado['valor_basico'].apply(
                    lambda v: '{:.2f}'.format(float(v)).replace('.', ','))
            else:
                resultado = resultado.iloc[:, [2, 6, 7, 8, 9]]
                resultado.columns = ['referencia', 'valor', 'dci', 'dij', 'dcm']
                resultado['valor_basico'] = resultado['valor'].apply(
                    lambda v: '{:.2f}'.format(float(v)).replace('.', ','))
            resultado['valor'] = resultado['valor'].apply(
                lambda v: '{:.2f}'.format(float(v)).replace('.', ','))
            resultado['dci'] = resultado['dci'].apply(
                lambda d: None if isinstance(d, np.float) and np.isnan(d) else f'{d.strftime("%d/%m/%y")}'
            )
            resultado['dij'] = resultado['dij'].apply(
                lambda d: None if isinstance(d, np.float) and np.isnan(d) else f'{d.strftime("%d/%m/%y")}'
            )
            resultado['dcm'] = resultado['dcm'].apply(
                lambda d: f'{d.strftime("%d/%m/%y")}'
            )
            if alinea in ['h', 'i', 'j']:
                resultado['davb'] = resultado['dij']
            else:
                raise Exception(f'Não sei ainda tratar o DAVB para planilha de glosa de créditos para a alinea {alinea}!')
            return resultado
        except Exception as e:
            logger.exception('Ocorreu uma falha ao copiar a planilha de créditos '
                             'template para a pasta da fiscalização')
            raise e
        finally:
            if wb is not None:
                wb.close()

    def get_sheet_names(self) -> list:
        return self.workbook().sheetnames

    def conta_fiscal_path(self) -> Path:
        return self.main_path / 'Conta Fiscal.pdf'

    def get_vencimentos_GIA(self) -> pd.DataFrame:
        audit = Audit.get_current_audit()
        refresh_cf = False
        if not self.conta_fiscal_path().is_file():
            refresh_cf = True
        else:
            vencimentos = PDFExtractor.vencimentos_de_PDF_CFICMS(self.conta_fiscal_path(), self.main_path,
                                                                 audit.get_periodos_da_fiscalizacao(rpa=True))
            if GeneralFunctions.last_day_of_month(vencimentos['vencimento'].iloc[-1]) < \
                    GeneralFunctions.first_day_of_month_before(datetime.date.today()) \
                    and vencimentos['saldo'].iloc[-1] < 0:
                refresh_cf = True
        if refresh_cf:
            with SeleniumWebScraper(self.main_path) as ws:
                ws.get_conta_fiscal(audit.ie, audit.inicio_auditoria.year, audit.fim_auditoria.year,
                                    self.conta_fiscal_path())
                self.conta_fiscal_path().unlink(missing_ok=True)
                self.creditos_path.unlink(missing_ok=True)
            vencimentos = PDFExtractor.vencimentos_de_PDF_CFICMS(self.conta_fiscal_path(), self.main_path,
                                                                 audit.get_periodos_da_fiscalizacao(rpa=True))
        return vencimentos

    def update_number_in_subtotals(self, sheet_name: str, item: int):
        ws = self.workbook()[sheet_name]
        logger.info(f'Atualizando número dos itens e subitens da aba {sheet_name}...')
        subitem = 1
        for row in ws.iter_rows(min_row=self._get_template_row() + 1, max_row=ws.max_row, max_col=1):
            for cell in row:
                if isinstance(cell.value, str):
                    if cell.value.find('Total Subitem') >= 0:
                        cell.value = f'Total Subitem {item}.{subitem}'
                        subitem += 1
                    elif cell.value.find('TOTAL ITEM') >= 0:
                        cell.value = f'TOTAL ITEM {item}'
        self.salva_planilha()

    def get_ddf_from_sheet(self, sheet_name: str, infraction: Infraction, item: int = 1):
        sheet = self.planilha(sheet_name)
        is_monthly_grouped = len(list(filter(lambda k: k.upper() == 'MÊS', sheet.keys()))) > 0
        titulos_planilha = sheet.keys().tolist()
        if not is_monthly_grouped and any([titulo in titulos_planilha for titulo in ['Referência', 'Período']]):
            # se a planilha não é agrupada, vou supor que última coluna tem os valores
            titulo = [titulo for titulo in titulos_planilha if titulo in ['Referência', 'Período']][0]
            aba = sheet.dropna(axis=0, how='all')
            aba['item'] = aba.index
            aba = aba.iloc[:, [-1, -2, titulos_planilha.index(titulo)]]
            aba.columns = ['item', 'valor', 'referencia']
            aba['referencia'] = aba['referencia'].apply(lambda r: GeneralFunctions.last_day_of_month(r))
        else:
            aba = sheet[
                sheet.iloc[:, 0].map(lambda texto: isinstance(texto, str) and bool(re.search('Subitem', texto)))]
            aba = aba.dropna(axis=1)
            if aba.empty:
                # pode ser que seja um caso de não totalização - cada linha é uma referência
                raise ExcelArrazoadoIncompletoException(
                    f'Planilha {sheet_name} não tem coluna "Referência" ou "Período" '
                    f'(quando itens já estão agrupados) ou ela tem vários itens '
                    f'da mesma referência, mas está sem totalizadores!\n'
                    f'No segundo caso, execute a macro com CTRL+SHIFT+E, '
                    f'salve, feche a planilha e tente novamente!')
            if len(aba.columns) == 4:
                aba.columns = ['item', 'valor_basico', 'valor', 'referencia']
            elif len(aba.columns) == 3:
                aba.columns = ['item', 'valor', 'referencia']
            else:
                raise Exception(f'Dataframe da planilha {sheet_name} tem colunas diferente do esperado!')

        aba['referencia'] = aba['referencia'].astype('datetime64[ns]')

        if infraction.inciso == 'I' and infraction.alinea == 'a':
            aba['valor'] = aba['valor'].apply(lambda v: '{:.2f}'.format(float(v)).replace('.', ','))
            ddf = aba[['valor', 'referencia']]
            ddf['dia_seguinte'] = ddf['referencia'].apply(
                lambda d: f'{(d + datetime.timedelta(days=1)).strftime("%d/%m/%y")}'
            )
        elif infraction.inciso == 'I' and infraction.alinea in ['b', 'c', 'd', 'i', 'j', 'l']:
            if is_monthly_grouped:
                vencimentos = self.get_vencimentos_GIA()
                aba['valor'] = aba['valor'].apply(lambda v: '{:.2f}'.format(float(v)).replace('.', ','))
                ddf = aba.merge(vencimentos, how='left', on='referencia')
                if len(ddf.dropna(axis=0)) < len(ddf):
                    raise ExcelArrazoadoIncompletoException(
                        'Não foram encontradas datas de vencimento no PDF da Conta Fiscal ICMS '
                        'para todas as referências da planilha! '
                        'Verifique se o PDF precisa ser complementado e apague o arquivo cficms.json na pasta Dados!'
                    )
                ddf['vencimento'] = ddf['vencimento'].apply(
                    lambda d: f'{(d + datetime.timedelta(days=1)).strftime("%d/%m/%y")}'
                )
                ddf = ddf[['valor', 'referencia', 'vencimento']]
            else:
                aba['valor'] = aba['valor'].apply(lambda v: '{:.2f}'.format(float(v)).replace('.', ','))
                ddf = aba[['valor', 'referencia']]
                ddf['vencimento'] = ddf['referencia'].apply(
                    lambda d: f'{(d + datetime.timedelta(days=1)).strftime("%d/%m/%y")}'
                )
        elif infraction.inciso == 'I' and infraction.alinea == 'e':
            aba['valor'] = aba['valor'].apply(lambda v: '{:.2f}'.format(float(v)).replace('.', ','))
            ddf = aba[['valor', 'referencia']]
            ddf['davb'] = ddf['referencia'].apply(
                lambda d: f'{(d + datetime.timedelta(days=1)).strftime("%d/%m/%y")}'
            )
        elif infraction.inciso == 'II':
            if infraction.alinea != 'j' and len(aba.columns) < 4:
                raise ExcelArrazoadoIncompletoException(
                    'Planilha precisa ter 2 subtotais: valor básico e imposto, nesta ordem. '
                    'Refaça a totalização da planilha selecionando as 2 colunas simultaneamente, '
                    'antes de rodar a macro com CTRL+SHIFT+E')
            vencimentos = self.get_vencimentos_GIA()
            self.create_creditos_sheet(vencimentos)
            ddf = aba.merge(vencimentos, on='referencia', how='right')
            ddf = self.generate_creditos_ddf(ddf, infraction.alinea, item)
        elif infraction.inciso == 'IV' and infraction.alinea == 'b':
            aba['valor'] = aba['valor'].apply(lambda v: '{:.2f}'.format(float(v)).replace('.', ','))
            ddf = aba[['referencia', 'valor']]
        elif infraction.inciso == 'IV' and infraction.alinea in ['z2', 'z3']:
            aba['valor'] = aba['valor'].apply(lambda v: f'{v}')
            ddf = aba[['referencia', 'valor']]
        elif infraction.inciso == 'V' and infraction.alinea in ['a', 'c', 'm']:
            aba['valor'] = aba['valor'].apply(lambda v: '{:.2f}'.format(float(v)).replace('.', ','))
            ddf = aba[['referencia', 'valor']]
        elif infraction.inciso == 'VII' and infraction.alinea == 'a':
            if len(aba.columns) == 4:
                aba.columns = ['referencia', 'entrega', 'atraso', 'valor']
                ddf = aba[['referencia', 'valor', 'atraso']]
            else:
                if len(aba.columns) == 2:
                    aba.columns = ['referencia', 'valor']
                ddf = aba[['referencia', 'valor']]
                ddf['atraso'] = None
        else:
            raise ExcelArrazoadoCriticalException(f'Inciso/alinea não mapeados: '
                                                  f'{infraction.inciso}, {infraction.alinea}')

        assert type(ddf) == pd.DataFrame, \
            'Não gerou uma listagem: não deve ter dados na planilha, ou os dados não foram bem filtrados'
        assert len(ddf) > 0, 'Listagem está vazia, deu algum problema na leitura dos dados'

        ddf['referencia'] = ddf['referencia'].apply(
            lambda d: f'{d.strftime("%d/%m/%y")}'
        )

        return {'infracao': infraction, 'ddf': ddf, 'mensal': is_monthly_grouped}

    def generate_dados_efds_e_imprime(self, path: Path):
        logger.info('Gerando listagem de EFDs...')
        ws = self.workbook()['EFDs']
        try:
            if ws.max_row <= 10:
                with SQLReader(Audit.get_current_audit().schema) as postgres:
                    if not postgres.does_table_exist('escrituracaofiscal'):
                        return
                    qtd, df = postgres.executa_consulta(
                        "SELECT replace(to_char(cpf_cnpj::bigint, '00:000:000/0000-00'), ':', '.') as \"CNPJ\", "
                        "REPLACE(to_char(ie::bigint, '000:000:000:000'), ':', '.') AS \"IE\", "
                        "efd.nome_contribuinte as \"Contribuinte\", "
                        "to_char(efd.dataInicial, 'mm/yyyy') as \"Referência\", "
                        "to_date(substring(efd.localizacaoarquivo FROM '(\d{8})\d+.txt'), 'ddmmyyyy') "
                        "AS \"Data Recepção\", "
                        "efd.hasharquivo::varchar AS \"Hash\" "
                        "FROM escrituracaofiscal as efd ORDER BY efd.datainicial")
                for linha in _dataframe_to_rows(df):
                    ws.append(linha)
                # coloca bordas nos dados
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                for row in ws.iter_rows(min_row=11, max_row=ws.max_row, max_col=5):
                    for cell in row:
                        cell.border = thin_border
                self.salva_planilha()
            self.imprime_planilha('EFDs', path)
        except Exception as e:
            logger.exception('Erro no preenchimento de dados de EFD na planilha de arrazoado')
            raise e

    def periodos_de_referencia(self, sheet_name, freq='M'):
        aba = self.planilha(sheet_name)
        if type(aba.iloc[-2, -1]) in (datetime.datetime, datetime.date):
            aba = aba.iloc[:-1, -1].dropna(axis=0)
        else:
            # considera que a última coluna de data (não timestamp preenchido) encontrada é a referência
            colunas_data = [idx for idx, vlw in enumerate(aba.iloc[-1].tolist())
                            if isinstance(vlw, datetime.date) or isinstance(vlw, datetime.datetime)]
            colunas_data.reverse()
            achou = False
            for c in colunas_data:
                data_types = aba.iloc[:, 1].map(lambda x: type(x)).unique().tolist()
                if len(data_types) == 1 and data_types[0] == datetime.date:
                    aba = aba.iloc[:, c]
                    achou = True
                    break
                datas = aba.iloc[:, c].map(lambda x: x.time())
                if len(datas[datas != datetime.time()]) == 0:
                    aba = aba.iloc[:, c]
                    achou = True
                    break
            if not achou:
                raise ExcelArrazoadoIncompletoException(f'Planilha {sheet_name} não '
                                                        f'tem uma coluna de data sem horário!')
        if freq == 'M':
            retorno = [GeneralFunctions.last_day_of_month(r) for r in aba]
        else:
            retorno = [datetime.date(r.year, 1, 1) for r in aba]
        retorno = list(set(retorno))
        retorno.sort()
        return retorno

    def modelos_documento_fiscal(self, sheet_name):
        try:
            aba = self.planilha(sheet_name)
        except KeyError:
            raise ExcelArrazoadoAbaInexistenteException(f'Aba {sheet_name} não existe mais na planilha!')
        colunas_modelo = [nome for nome in aba.keys().tolist() if 'MODELO' in nome.upper()]
        if len(colunas_modelo) == 0:
            colunas_chave = [nome for nome in aba.keys().tolist() if 'CHAVE' in nome.upper()]
            if len(colunas_chave) > 0:
                aba = aba[aba[colunas_chave[0]].str.get(0) <= '9']
                modelos = pd.to_numeric(aba[colunas_chave[0]].str[20:22])
            else:
                raise Exception(f'Não achei coluna na aba {sheet_name} com um campo contendo "Modelo" ou "Chave"')
        else:
            modelos = aba[colunas_modelo[0]][aba[colunas_modelo[0]].map(lambda v: not isinstance(v, str) and
                                                                                  np.isfinite(v) and
                                                                                  round(v - int(v), 2) == 0)]
        return modelos.sort_values().unique()

    def imprime_planilha(self, sheet, report_full_path: Path, path: Path = None, item: int = None) -> Path:
        if item:
            # coloca número do anexo na planilha e mostra detalhes, caso subtotais estejam fechados
            # self.update_number_in_subtotals(sheet, item)
            openpyxlws = self.workbook()[sheet]
            is_grouped = any(x > 0 for x in
                             [v.outline_level for v in [openpyxlws.row_dimensions[idx]
                                                        for idx in range(self._get_template_row(),
                                                                         openpyxlws.max_row)]])
            if is_grouped:
                for idx in range(self._get_template_row(), openpyxlws.max_row):
                    openpyxlws.row_dimensions[idx].hidden = False
            self.salva_planilha()

        wb_path = self.planilha_path if path is None else path
        if type(sheet) == str:
            sheet_number = self.get_sheet_names().index(sheet)
        else:
            sheet_number = sheet
        return print_workbook_as_pdf(wb_path, report_full_path, sheet_number=sheet_number)

    def salva_planilha(self, wb: Workbook = None, path: Path = None):
        caminho = self.planilha_path if path is None else path
        workbook = self.workbook() if wb is None else wb
        try:
            workbook.save(caminho)
        except PermissionError as pe:
            raise ExcelArrazoadoCriticalException(
                f'Arquivo Excel {caminho} está aberto, feche-o e tente novamente.')
        finally:
            workbook.close()
            if wb is None:
                del self._workbook
                self._workbook = None
        # refresh via com, para salvar os valores calculados
        refresh_planilha(caminho)

    def exporta_relatorio_para_planilha(self, sheet_name: str, analysis: Analysis, df: pd.DataFrame):
        self._exporta_relatorio_para_planilha(sheet_name, analysis, df)
        if len(analysis.filter_columns()) > 0:
            for infraction in analysis.infractions:
                self._exporta_relatorio_para_planilha(sheet_name, analysis, df, infraction=infraction)
        self.salva_planilha()

    def _exporta_relatorio_para_planilha(self, sheet_name: str, analysis: Analysis, df: pd.DataFrame,
                                         infraction: Infraction = None):
        ws = None
        try:
            ws = self.workbook().copy_worksheet(self.workbook()['template'])
            if infraction and infraction.has_filtro():
                ws.title = infraction.sheet_extended_name(sheet_name)
            else:
                ws.title = sheet_name
            celula_cabecalho_template = ws[f'A{ws.max_row}']
            template_row = self._get_template_row()
            columns_to_drop = []
            if infraction:
                columns_to_drop = infraction.analysis.filter_columns()
                columns_to_drop.remove(infraction.filtro_coluna)

            cabecalhos = [c for c in df.keys().tolist() if c not in columns_to_drop]

            ws['A7'] = analysis.name if infraction is None else infraction.planilha_titulo
            # faz merge das primeiras linhas, até a última linha antes do cabeçalho
            coluna_A = [row[0] for row in ws[f'A{ws.min_row}:A{ws.max_row}']]
            coluna_A.reverse()
            for cell in coluna_A:
                if cell.fill.bgColor.value == '00000000':
                    ws.merge_cells(start_row=cell.row, end_row=cell.row,
                                   start_column=1, end_column=len(cabecalhos))

            # adiciona a imagem do brasao na planilha
            img = Image(r'resources/brasao.png')
            img.anchor = 'A1'
            # fazendo resize da imagem na mao
            img.height *= 0.556
            img.width *= 0.556
            ws.add_image(img)

            # faz um subconjunto do dataframe, caso tenha filtro
            if infraction and infraction.has_filtro():
                df = df.drop(labels=columns_to_drop, axis=1)
                if infraction.is_positive_filter():
                    df = df[df[infraction.filtro_coluna] >= 0]
                else:
                    df = df[df[infraction.filtro_coluna] <= 0]
                if len(df) == 0:
                    self.workbook().remove(ws)
                    return

            # joga valores do dataframe na planilha
            for linha in _dataframe_to_rows(df):
                ws.append(linha)

            # formata valores
            df = df.convert_dtypes()
            for idx in range(1, len(df.dtypes) + 1):
                col_type = df.dtypes.tolist()[idx - 1]
                if col_type == 'datetime64[ns]':
                    for row in ws.iter_rows(min_row=template_row + 1, max_row=ws.max_row, min_col=idx, max_col=idx):
                        if cabecalhos[idx - 1] in ('Período', 'Referência'):
                            row[0].number_format = 'mm/yyyy'
                        else:
                            row[0].number_format = 'dd/mm/yyyy'
                        row[0].alignment = Alignment(horizontal='center', vertical='center')
                elif col_type == 'Float64':
                    for row in ws.iter_rows(min_row=template_row + 1, max_row=ws.max_row, min_col=idx, max_col=idx):
                        row[0].number_format = r'_-\R$ * #,##0.00_-'
                elif col_type == 'Int64':
                    for row in ws.iter_rows(min_row=template_row + 1, max_row=ws.max_row, min_col=idx, max_col=idx):
                        row[0].number_format = '@'
                        row[0].alignment = Alignment(horizontal='center', vertical='center')

            # entende que o formato do cabeçalho da tabela é igual da última linha da coluna A do template
            # como fez append do dataframe, a linha template precisa ser removida depois
            for i in range(1, len(cabecalhos) + 1):
                ws.cell(row=template_row + 1, column=i)._style = copy(celula_cabecalho_template._style)
            ws.delete_rows(template_row)

            # coloca bordas nos dados
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            for row in ws.iter_rows(min_row=template_row + 1, max_row=ws.max_row, max_col=len(cabecalhos)):
                for cell in row:
                    cell.border = thin_border

            # tenta fazer autosize das colunas
            for column_cells in ws.columns:
                if column_cells[-1].number_format.endswith('yyyy'):
                    length = max(len(column_cells[template_row - 1].value), len(column_cells[-1].number_format)) + 1
                else:
                    length = max(len(str(cell.value) or "") + 1 for cell in column_cells[template_row - 1:])

                if column_cells[-1].number_format == r'_-\R$ * #,##0.00_-':
                    # adiciona os caracteres de formatacao de moeda
                    length += int((length - 2) / 3) + 4
                logger.debug(f'Coluna {column_cells[-1].column_letter} terá largura {length}')
                ws.column_dimensions[column_cells[-1].column_letter].width = length

            # opcoes de impressao
            ws.print_options.horizontalCentered = True
            ws.print_title_rows = f'{template_row}:{template_row}'
            ws.page_margins.top = 1
            ws.page_margins.bottom = 1
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.header = 0.5
            ws.page_margins.footer = 0.5
            if len(cabecalhos) >= 10:
                ws.page_setup.orientation = 'landscape'
            else:
                ws.page_setup.orientation = 'portrait'
            ws.page_setup.fitToWidth = True
            ws.evenFooter.center.text = "Página &[Page] de &N"
            ws.oddFooter.center.text = "Página &[Page] de &N"
        except Exception as e:
            logger.exception('Falha na exportação de tabela para Excel')
            if ws:
                self.workbook().remove(ws)
            raise ExcelArrazoadoCriticalException(str(e))

    def gera_quadro_3(self, quadro_1: Path):
        logger.info('Extraindo informações do Quadro 1 do AIIM para gerar Quadro 3...')
        df = PDFExtractor.get_quadro_1_data(quadro_1)
        wb = None
        logger.info('Gerando arquivo Quadro 3.xlsm...')
        try:
            shutil.copyfile(Path('resources') / 'Quadro 3.xlsm', self.quadro3_path)
            wb = openpyxl.load_workbook(self.quadro3_path, keep_vba=True)
            ws = wb['DDF Original']
            ws['C2'].value = Audit.get_current_audit().cnpj
            ws['G2'].value = Audit.get_current_audit().empresa
            ws['C4'].value = Audit.get_current_audit().aiim_number
            with MDBReader.MDBReader() as aiim2003:
                ws['Q4'].value = aiim2003.get_last_ufesp_stored()
            ws_row = 9
            for linha in df:
                for col, val in enumerate(linha, start=2):
                    # é necessário pular uma coluna a partir da 13, pelo formato da planilha
                    ws.cell(row=ws_row, column=(col + 1 if col >= 13 else col), value=val)
                ws_row += 1
            self.salva_planilha(wb, self.quadro3_path)
            logger.info('Gerando Quadro 3.pdf...')
            self.imprime_planilha(4, self.main_path / 'AIIM' / 'Quadro 3.pdf', path=self.quadro3_path)
        except Exception as e:
            logger.exception('Falha no preenchimento do Quadro 3')
            self.quadro3_path.unlink(missing_ok=True)
            raise ExcelArrazoadoCriticalException(f'Falha no preenchimento do Quadro 3: {str(e)}')
        finally:
            if wb:
                wb.close()

    def get_operations_for_aiim(self, operacoes_xls: Path):
        csv_df = None
        wb = None
        if self.operacoes_csv.is_file():
            csv_df = pd.read_csv(str(self.operacoes_csv), index_col=0, parse_dates=True)
        if not operacoes_xls.is_file():
            if csv_df is not None:
                return csv_df
            else:
                return pd.DataFrame()
        try:
            logger.info('Encontrei arquivo de relatório Valor Total Documentos Fiscais x GIA, processando valores...')
            wb = openpyxl.load_workbook(str(operacoes_xls))
            ws = wb['Doc Fiscais x GIA']
            data = []
            refs = []

            # usa a coluna de documentos se for maior que a de GIA, e vice-versa
            # não mistura os valores pra evitar problemas na defesa do AIIM
            row = 5
            while type(ws.cell(row=row, column=1).value) == int:
                row += 1
            if type(ws.cell(row=row, column=8).value) == float:
                total_docs = ws.cell(row=row, column=8).value
            else:
                valor = ws.cell(row=row, column=8).value
                if valor:
                    total_docs = float(str(valor).replace('.', '').replace(',', '.'))
                else:
                    total_docs = 0
            row = 5
            while type(ws.cell(row=row, column=10).value) == int:
                row += 1
            if type(ws.cell(row=row, column=11).value) == float:
                total_gias = ws.cell(row=row, column=11).value
            else:
                valor = ws.cell(row=row, column=11).value
                if valor:
                    total_gias = float(str(valor).replace('.', '').replace(',', '.'))
                else:
                    total_gias = 0
            if total_docs > total_gias:
                logger.info('Decidido usar tabela de documentos fiscais por ter valores maiores, para cadastro de VTO')
                coluna_ref = 1
                coluna_valor = 8
            else:
                logger.info('Decidido usar tabela de GIA por ter valores maiores, para cadastro de VTO')
                coluna_ref = 10
                coluna_valor = 11

            row = 5
            while type(ws.cell(row=row, column=coluna_ref).value) == int:
                ref = ws.cell(row=row, column=coluna_ref).value
                valor = ws.cell(row=row, column=coluna_valor).value
                periodo = datetime.datetime.strptime(str(ref), '%Y%m').date()
                data.append([int(str(ref)[4:]), int(str(ref)[:4]), valor])
                refs.append(periodo)
                row += 1
            new_df = pd.DataFrame(data, columns=['Mes', 'Ano', 'Valor Contabil - CFOP'], index=pd.to_datetime(refs))
            refs.sort()
            if self.operacoes_csv.is_file() and refs[0] < GeneralFunctions.first_day_of_month_before(
                    datetime.date.today()):
                ops_df = pd.concat([csv_df, new_df])
            else:
                ops_df = new_df
            ops_df = ops_df.drop_duplicates(subset=['Ano', 'Mes'], keep='last').sort_values(
                by=['Ano', 'Mes'], ascending=False)
            ops_df.to_csv(self.operacoes_csv)
            if operacoes_xls.is_file():
                operacoes_xls.unlink()
            return ops_df
        finally:
            if wb:
                wb.close()

    def clear_cache(self):
        self._planilha = {}
        if self._workbook is not None:
            self._workbook.close()
            del self._workbook
            self._workbook = None
